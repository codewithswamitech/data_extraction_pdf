#!/usr/bin/env python3
"""
PDF to Excel Table Extractor — Hybrid Engine

Combines pdfplumber's native PDF text extraction with automatic detection
of sub-column positions. This handles PDFs where sub-columns (like
LIFE, PENSION, HEALTH, VAR.INS) share a single bordered cell without
any vertical lines between them.

Approach:
  1. Extract all words with exact (x,y) positions from the PDF
  2. Detect ALL vertical edges (from lines + rectangles in PDF)
  3. Also detect text-based column positions by analyzing word x-positions
  4. Merge both to create comprehensive column boundaries
  5. Extract tables using these explicit vertical lines

Usage:
    python extract_tables_hybrid.py <pdf_file> [output_excel]
"""

import sys
import os
import time
from pathlib import Path
from collections import Counter

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("  pip install pdfplumber openpyxl")
    sys.exit(1)


# ── Styling ───────────────────────────────────────────────────────
HEADER_BG = '1F4E79'
HEADER_FG = 'FFFFFF'
EVEN_BG   = 'FFFFFF'
ODD_BG    = 'F5F5F5'


def get_pdf_vertical_edges(page):
    """Get all vertical edge x-positions from PDF vector elements."""
    edges = set()
    # From explicit lines
    for l in page.lines:
        if abs(l['x0'] - l['x1']) < 2:  # vertical line
            edges.add(round(l['x0'], 0))
    # From rectangle boundaries
    for r in page.rects:
        edges.add(round(r['x0'], 0))
        edges.add(round(r['x1'], 0))
    return sorted(edges)


def get_word_column_positions(page, pdf_edges):
    """
    Analyze word x-positions to find sub-column boundaries
    that aren't represented by drawn lines in the PDF.

    Returns a sorted list of all vertical boundary x-positions:
    PDF edges + text-derived boundaries.
    """
    words = page.extract_words(x_tolerance=2, y_tolerance=2)
    if not words:
        return pdf_edges

    # Collect left-edge positions of ALL words
    all_x0 = [round(w['x0'], 0) for w in words]

    # Cluster nearby x-positions (within 6 pts = same column start)
    all_x0.sort()
    clusters = []
    current = [all_x0[0]]
    for x in all_x0[1:]:
        if x - current[-1] <= 6:
            current.append(x)
        else:
            clusters.append((sum(current) / len(current), len(current)))
            current = [x]
    if current:
        clusters.append((sum(current) / len(current), len(current)))

    # Only keep clusters that have enough words (appear in multiple rows)
    # This prevents stray text from creating false columns
    min_count = max(3, len(words) * 0.01)  # At least 1% of total words
    strong_columns = [pos for pos, count in clusters if count >= min_count]

    # Merge with PDF edges
    all_boundaries = set()
    for e in pdf_edges:
        all_boundaries.add(e)

    # Add text-derived boundaries, but only if they're NOT close to
    # an existing PDF edge (to avoid duplicates)
    for col_x in strong_columns:
        is_near_edge = any(abs(col_x - e) < 10 for e in pdf_edges)
        if not is_near_edge:
            all_boundaries.add(col_x - 3)  # Place line just before column start

    return sorted(all_boundaries)


def extract_table_with_explicit_lines(page, v_lines):
    """
    Extract tables using explicit vertical lines for column detection.
    Horizontal strategy stays as 'text' for proper row splitting.
    """
    settings = {
        "vertical_strategy": "explicit",
        "explicit_vertical_lines": v_lines,
        "horizontal_strategy": "text",
        "snap_tolerance": 4,
        "join_tolerance": 4,
        "edge_min_length": 8,
        "min_words_vertical": 2,
        "intersection_y_tolerance": 10,
    }
    return page.extract_tables(settings)


def extract_tables_from_pdf(pdf_path):
    """Extract all tables from PDF using hybrid line + text detection."""
    print(f"Reading: {pdf_path}")
    all_tables = []

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        print(f"Pages: {total}")
        print()

        for page_num, page in enumerate(pdf.pages, start=1):
            # Get PDF vector edges
            pdf_edges = get_pdf_vertical_edges(page)

            # Get combined boundaries (PDF edges + text-derived)
            all_v_lines = get_word_column_positions(page, pdf_edges)

            if len(all_v_lines) < 2:
                continue

            # Extract tables with explicit vertical lines
            tables = extract_table_with_explicit_lines(page, all_v_lines)

            if tables:
                for idx, table in enumerate(tables, start=1):
                    if table and len(table) > 0:
                        # Filter out completely empty rows
                        clean = [r for r in table
                                 if any(c and str(c).strip() for c in r)]
                        if clean:
                            all_tables.append({
                                'table': clean,
                                'page': page_num,
                                'index': idx,
                                'cols': max(len(r) for r in clean),
                            })

            # Progress indicator every 20 pages
            if page_num % 20 == 0 or page_num == total:
                print(f"  Processed {page_num}/{total} pages "
                      f"({len(all_tables)} tables so far)")

    return all_tables


def create_excel(tables, output_path):
    """Create styled Excel workbook from extracted tables."""
    wb = Workbook()
    wb.remove(wb.active)

    hdr_font  = Font(name='Times New Roman', size=11, bold=True, color=HEADER_FG)
    hdr_fill  = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cel_font  = Font(name='Times New Roman', size=10, color='000000')
    cel_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    wh_fill   = PatternFill(start_color=EVEN_BG, end_color=EVEN_BG, fill_type='solid')
    gr_fill   = PatternFill(start_color=ODD_BG, end_color=ODD_BG, fill_type='solid')
    border    = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for t_idx, td in enumerate(tables, start=1):
        table = td['table']
        ws = wb.create_sheet(title=f'Table_{t_idx}')

        # Write data
        for ri, row in enumerate(table, start=1):
            for ci, val in enumerate(row, start=1):
                if val is not None:
                    v = str(val).strip() if isinstance(val, str) else val
                    ws.cell(row=ri, column=ci, value=v)

        mr = ws.max_row or 1
        mc = ws.max_column or 1

        # Apply styles
        for r in range(1, mr + 1):
            for c in range(1, mc + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if r == 1:
                    cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, hdr_align
                else:
                    cell.font = cel_font
                    cell.alignment = cel_align
                    cell.fill = gr_fill if r % 2 == 0 else wh_fill

        # Column widths
        for c in range(1, mc + 1):
            mx = 0
            for r in range(1, mr + 1):
                v = ws.cell(row=r, column=c).value
                if v:
                    mx = max(mx, len(str(v).replace('\n', ' ')))
            ws.column_dimensions[get_column_letter(c)].width = max(min(mx + 2, 50), 10)

        if mr > 1:
            ws.freeze_panes = 'A2'

        print(f"  Table_{t_idx}: {mr} rows × {mc} cols (page {td['page']})")

    # Save
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_path)
    size_mb = os.path.getsize(output_path) / (1024 * 1024)
    print(f"\n✓ Saved: {output_path} ({size_mb:.2f} MB, {len(tables)} tables)")


def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_tables_hybrid.py <pdf> [output.xlsx]")
        sys.exit(1)

    pdf_path = sys.argv[1]
    if not os.path.exists(pdf_path):
        print(f"Error: Not found: {pdf_path}")
        sys.exit(1)

    if len(sys.argv) >= 3:
        out = sys.argv[2]
    else:
        name = Path(pdf_path).stem
        out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Output_excel')
        os.makedirs(out_dir, exist_ok=True)
        out = os.path.join(out_dir, f"{name}_hybrid.xlsx")

    if not out.lower().endswith('.xlsx'):
        out += '.xlsx'

    print("=" * 60)
    print("PDF to Excel — Hybrid Engine")
    print("(PDF lines + text-position column detection)")
    print("=" * 60)
    print()

    start = time.time()
    tables = extract_tables_from_pdf(pdf_path)

    if not tables:
        print("\nNo tables found.")
        sys.exit(0)

    print(f"\nTotal: {len(tables)} tables")
    print()
    create_excel(tables, out)
    print(f"Time: {time.time()-start:.1f}s")


if __name__ == "__main__":
    main()

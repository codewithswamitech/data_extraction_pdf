#!/usr/bin/env python3
"""
PDF to Excel Table Extractor - With Merged Column Support

This version properly handles tables with hierarchical/merged column headers
where multiple sub-columns are grouped under one main column.
"""

import sys
import os
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Error: Required library not installed: {e}")
    print("\nPlease install the required libraries:")
    print("  pip install pdfplumber openpyxl")
    sys.exit(1)


# Color scheme constants
HEADER_BG_COLOR = '1F4E79'  # Dark blue
HEADER_TEXT_COLOR = 'FFFFFF'  # White
ROW_EVEN_COLOR = 'FFFFFF'    # White
ROW_ODD_COLOR = 'F5F5F5'     # Light gray
BORDER_COLOR = '000000'      # Black


def create_styles():
    """Create and return style objects for Excel formatting."""
    
    # Header style
    header_font = Font(
        name='Times New Roman',
        size=11,
        bold=True,
        color=HEADER_TEXT_COLOR
    )
    header_fill = PatternFill(
        start_color=HEADER_BG_COLOR,
        end_color=HEADER_BG_COLOR,
        fill_type='solid'
    )
    header_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Cell style (data rows)
    cell_font = Font(
        name='Times New Roman',
        size=10,
        color='000000'
    )
    cell_alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True
    )
    
    # Fill colors
    white_fill = PatternFill(
        start_color=ROW_EVEN_COLOR,
        end_color=ROW_EVEN_COLOR,
        fill_type='solid'
    )
    gray_fill = PatternFill(
        start_color=ROW_ODD_COLOR,
        end_color=ROW_ODD_COLOR,
        fill_type='solid'
    )
    
    # Border style
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_alignment': header_alignment,
        'cell_font': cell_font,
        'cell_alignment': cell_alignment,
        'white_fill': white_fill,
        'gray_fill': gray_fill,
        'border': thin_border
    }


def detect_header_rows(table):
    """
    Detect which rows are header rows and identify column groupings.
    
    Returns:
        tuple: (header_row_count, column_groups)
        where column_groups is a list of (start_col, end_col, group_name)
    """
    if not table or len(table) < 2:
        return 1, []
    
    # Heuristics to detect header rows
    header_row_count = 1
    column_groups = []
    
    # Look for patterns that indicate grouped headers
    for row_idx in range(min(3, len(table))):
        row = table[row_idx]
        if not row:
            continue
        
        # Count non-empty cells
        non_empty = sum(1 for cell in row if cell and str(cell).strip())
        
        # If this row has significantly fewer non-empty cells than the next row,
        # it might be a grouping row
        if row_idx + 1 < len(table):
            next_row = table[row_idx + 1]
            next_non_empty = sum(1 for cell in next_row if cell and str(cell).strip())
            
            if non_empty > 0 and next_non_empty > 0 and non_empty < next_non_empty:
                # This row likely contains grouped headers
                header_row_count = max(header_row_count, row_idx + 1)
                
                # Detect column groupings
                for col_idx, cell in enumerate(row):
                    if cell and str(cell).strip():
                        # Check how many consecutive cells in next row belong to this group
                        group_span = 1
                        next_col = col_idx + 1
                        while next_col < len(next_row):
                            if next_row[next_col] is None or not str(next_row[next_col]).strip():
                                group_span += 1
                                next_col += 1
                            else:
                                break
                        
                        if group_span > 1:
                            column_groups.append((col_idx, col_idx + group_span - 1, str(cell).strip()))
    
    return header_row_count, column_groups


def merge_cells_in_excel(ws, column_groups, header_row_count):
    """
    Merge cells in Excel for grouped column headers.
    
    Args:
        ws: Excel worksheet
        column_groups: List of (start_col, end_col, group_name) tuples
        header_row_count: Number of header rows
    """
    for start_col, end_col, group_name in column_groups:
        # Convert to 1-based Excel indices
        excel_start = start_col + 1
        excel_end = end_col + 1
        
        # Merge cells horizontally
        if excel_start < excel_end:
            # Merge cells across the header row(s)
            for row_num in range(1, header_row_count + 1):
                ws.merge_cells(
                    start_row=row_num,
                    start_column=excel_start,
                    end_row=row_num,
                    end_column=excel_end
                )


def extract_tables_from_pdf(pdf_path):
    """
    Extract all tables from a PDF file.
    
    Args:
        pdf_path (str): Path to the PDF file
        
    Returns:
        list: List of tables (each table is a list of rows)
    """
    print(f"Reading PDF file: {pdf_path}")
    
    all_tables = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f"Total pages in PDF: {len(pdf.pages)}")
            
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                
                if tables:
                    for idx, table in enumerate(tables, start=1):
                        all_tables.append({
                            'table': table,
                            'page': page_num,
                            'index_on_page': idx
                        })
                        
    except FileNotFoundError:
        print(f"Error: PDF file not found: {pdf_path}")
        sys.exit(1)
    except Exception as e:
        print(f"Error reading PDF file: {e}")
        sys.exit(1)
    
    return all_tables


def apply_header_styling(ws, max_row, max_col, styles):
    """
    Apply header styling to all header rows.
    
    Args:
        ws: Excel worksheet
        max_row: Maximum row number
        max_col: Maximum column number
        styles: Style dictionary
    """
    # Apply header styling to first row
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']


def create_excel_from_tables(tables, output_path):
    """
    Create an Excel workbook from extracted tables with merged column support.
    
    Args:
        tables (list): List of tables with metadata
        output_path (str): Path to save the Excel file
        
    Returns:
        str: Path to the created Excel file
    """
    print(f"\nCreating Excel workbook with merged column support...")
    
    # Create workbook and remove default sheet
    wb = Workbook()
    wb.remove(wb.active)
    
    # Get styles
    styles = create_styles()
    
    # Process each table
    for table_idx, table_data in enumerate(tables, start=1):
        table = table_data['table']
        page_num = table_data['page']
        
        # Create sheet name
        sheet_name = f'Table_{table_idx}'
        ws = wb.create_sheet(title=sheet_name)
        
        # Detect header rows and column groupings
        header_row_count, column_groups = detect_header_rows(table)
        
        print(f"  Table {table_idx}: {len(table)} rows, detected {header_row_count} header row(s), {len(column_groups)} column group(s)")
        
        # Write table data
        if table:
            for row_idx, row in enumerate(table, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value is not None:
                        # Handle None and empty values
                        if isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Apply formatting
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Merge cells for grouped columns
        if column_groups:
            merge_cells_in_excel(ws, column_groups, header_row_count)
            print(f"    Merged {len(column_groups)} column group(s)")
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                
                # Apply basic styles
                cell.font = styles['cell_font']
                cell.alignment = styles['cell_alignment']
                cell.border = styles['border']
                
                # Header row styling (first header_row_count rows)
                if row <= header_row_count:
                    cell.font = styles['header_font']
                    cell.fill = styles['header_fill']
                    cell.alignment = styles['header_alignment']
                # Alternating row colors (for data rows only)
                elif row > header_row_count:
                    if (row - header_row_count) % 2 == 0:
                        cell.fill = styles['gray_fill']
                    else:
                        cell.fill = styles['white_fill']
        
        # Auto-adjust column widths
        for col in range(1, max_col + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            # Check content in all rows for this column
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    if cell.value:
                        cell_str = str(cell.value)
                        cell_str = cell_str.replace('\n', ' ')
                        max_length = max(max_length, len(cell_str))
                except:
                    pass
            
            # Set column width (max 50, min 10)
            adjusted_width = min(max_length + 2, 50)
            adjusted_width = max(adjusted_width, 10)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header rows
        if max_row > header_row_count:
            ws.freeze_panes = f'A{header_row_count + 1}'
        
        print(f"  Created {sheet_name}: {max_row} rows × {max_col} columns (from page {page_num})")
    
    # Ensure download directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Save workbook
    try:
        wb.save(output_path)
        file_size = os.path.getsize(output_path)
        file_size_mb = file_size / (1024 * 1024)
        
        print(f"\n✓ Excel file created successfully!")
        print(f"  File: {output_path}")
        print(f"  Size: {file_size_mb:.2f} MB")
        print(f"  Total tables: {len(tables)}")
        
        return output_path
        
    except PermissionError:
        print(f"Error: Permission denied. Please close the Excel file if it's open: {output_path}")
        sys.exit(1)
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        sys.exit(1)


def main():
    """Main function to orchestrate the PDF to Excel conversion."""
    
    # Check command line arguments
    if len(sys.argv) < 2:
        print("PDF to Excel Table Extractor - With Merged Column Support")
        print("=" * 70)
        print("\nUsage:")
        print("  python extract_tables_with_merged_columns.py <pdf_file_path> [output_excel_path]")
        print("\nExamples:")
        print("  python extract_tables_with_merged_columns.py document.pdf")
        print("  python extract_tables_with_merged_columns.py document.pdf output.xlsx")
        print()
        sys.exit(1)
    
    pdf_path = sys.argv[1]
    
    # Validate PDF file
    if not os.path.exists(pdf_path):
        print(f"Error: File not found: {pdf_path}")
        sys.exit(1)
    
    if not pdf_path.lower().endswith('.pdf'):
        print("Warning: Input file does not have .pdf extension")
    
    # Determine output path
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        # Generate output filename from input filename
        pdf_name = Path(pdf_path).stem
        output_path = f"/home/z/my-project/download/{pdf_name}_Tables_Merged.xlsx"
    
    # Validate output path has .xlsx extension
    if not output_path.lower().endswith('.xlsx'):
        output_path += '.xlsx'
    
    print("=" * 70)
    print("PDF to Excel Table Extractor - With Merged Column Support")
    print("=" * 70)
    print()
    
    # Extract tables
    tables = extract_tables_from_pdf(pdf_path)
    
    if not tables:
        print("\nWarning: No tables found in the PDF file.")
        print("The PDF may not contain any tabular data.")
        sys.exit(0)
    
    print(f"\nTotal tables found: {len(tables)}")
    print()
    
    # Create Excel file
    create_excel_from_tables(tables, output_path)
    
    print()
    print("=" * 70)
    print("Process completed successfully!")
    print("=" * 70)
    print("\nThis version properly handles:")
    print("  ✓ Merged/hierarchical column headers")
    print("  ✓ Column groupings (e.g., 'Linked Business' → 'Life', 'Pension', 'Health')")
    print("  ✓ Proper cell merging in Excel")
    print("  ✓ Header row detection")


if __name__ == "__main__":
    main()

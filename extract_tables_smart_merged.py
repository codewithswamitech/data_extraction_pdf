#!/usr/bin/env python3
"""
PDF to Excel Table Extractor - Smart Merged Column Detection

This version uses advanced detection to properly handle tables with
hierarchical/merged column headers and preserves the exact structure.
"""

import sys
import os
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Error: Required library not installed: {e}")
    print("\nPlease install the required libraries:")
    print("  pip install pdfplumber openpyxl")
    sys.exit(1)


# Color scheme constants
HEADER_BG_COLOR = '1F4E79'
HEADER_TEXT_COLOR = 'FFFFFF'
ROW_EVEN_COLOR = 'FFFFFF'
ROW_ODD_COLOR = 'F5F5F5'
BORDER_COLOR = '000000'


def create_styles():
    """Create and return style objects for Excel formatting."""
    
    header_font = Font(
        name='Times New Roman',
        size=11,
        bold=True,
        color=HEADER_TEXT_COLOR
    )
    header_fill = PatternFill(
        start_color=HEADER_BG_COLOR,
        end_color=HEADER_BG_COLOR,
        fill_type='solid'
    )
    header_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    cell_font = Font(
        name='Times New Roman',
        size=10,
        color='000000'
    )
    cell_alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True
    )
    
    white_fill = PatternFill(
        start_color=ROW_EVEN_COLOR,
        end_color=ROW_EVEN_COLOR,
        fill_type='solid'
    )
    gray_fill = PatternFill(
        start_color=ROW_ODD_COLOR,
        end_color=ROW_ODD_COLOR,
        fill_type='solid'
    )
    
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_alignment': header_alignment,
        'cell_font': cell_font,
        'cell_alignment': cell_alignment,
        'white_fill': white_fill,
        'gray_fill': gray_fill,
        'border': thin_border
    }


def analyze_column_structure(table):
    """
    Analyze table structure to detect column groupings.
    
    Returns:
        dict: {
            'header_rows': int,
            'column_groups': list of (row, start_col, end_col, label)
        }
    """
    if not table or len(table) < 2:
        return {'header_rows': 1, 'column_groups': []}
    
    result = {
        'header_rows': 1,
        'column_groups': []
    }
    
    # Analyze first 3 rows for header patterns
    max_analysis_rows = min(3, len(table))
    
    for row_idx in range(max_analysis_rows):
        row = table[row_idx]
        if not row:
            continue
        
        # Find groups of consecutive empty cells
        # These often indicate merged cells above
        groups = []
        current_group = None
        
        for col_idx, cell in enumerate(row):
            cell_value = str(cell).strip() if cell else ''
            
            if cell_value:
                # This cell has content, check if it spans multiple columns
                # Look ahead to see if next cells are empty in the same row
                # but have content in rows below
                span = 1
                next_col = col_idx + 1
                
                while next_col < len(row):
                    next_cell_value = str(row[next_col]).strip() if row[next_col] else ''
                    
                    if not next_cell_value:  # Empty in current row
                        # Check if there's content in rows below
                        has_below_content = False
                        for below_row_idx in range(row_idx + 1, min(row_idx + 3, len(table))):
                            if below_row_idx < len(table):
                                below_cell = table[below_row_idx]
                                if next_col < len(below_cell) and below_cell[next_col]:
                                    has_below_content = True
                                    break
                        
                        if has_below_content:
                            span += 1
                            next_col += 1
                        else:
                            break
                    else:
                        break
                
                if span > 1:
                    # This is a merged header
                    result['column_groups'].append({
                        'row': row_idx + 1,  # 1-based
                        'start_col': col_idx + 1,
                        'end_col': col_idx + span,
                        'label': cell_value
                    })
                    
                    # Update header row count
                    result['header_rows'] = max(result['header_rows'], row_idx + 1)
    
    return result


def extract_tables_from_pdf(pdf_path):
    """Extract all tables from a PDF file."""
    print(f"Reading PDF file: {pdf_path}")
    
    all_tables = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f"Total pages in PDF: {len(pdf.pages)}")
            
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                
                if tables:
                    for idx, table in enumerate(tables, start=1):
                        all_tables.append({
                            'table': table,
                            'page': page_num,
                            'index_on_page': idx
                        })
                        
    except FileNotFoundError:
        print(f"Error: PDF file not found: {pdf_path}")
        sys.exit(1)
    except Exception as e:
        print(f"Error reading PDF file: {e}")
        sys.exit(1)
    
    return all_tables


def create_excel_from_tables(tables, output_path):
    """Create Excel workbook with proper merged column support."""
    
    print(f"\nCreating Excel workbook with smart merged column detection...")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    styles = create_styles()
    
    for table_idx, table_data in enumerate(tables, start=1):
        table = table_data['table']
        page_num = table_data['page']
        
        sheet_name = f'Table_{table_idx}'
        ws = wb.create_sheet(title=sheet_name)
        
        # Analyze column structure
        structure = analyze_column_structure(table)
        header_rows = structure['header_rows']
        column_groups = structure['column_groups']
        
        print(f"  Table {table_idx}: {len(table)} rows, {header_rows} header row(s), {len(column_groups)} merged cell(s)")
        
        # Write all data first
        if table:
            for row_idx, row in enumerate(table, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value is not None:
                        if isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Apply merged cells
        for group in column_groups:
            try:
                ws.merge_cells(
                    start_row=group['row'],
                    start_column=group['start_col'],
                    end_row=group['row'],
                    end_column=group['end_col']
                )
            except:
                pass  # Skip invalid merges
        
        # Apply formatting
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                
                cell.font = styles['cell_font']
                cell.alignment = styles['cell_alignment']
                cell.border = styles['border']
                
                # Header styling
                if row <= header_rows:
                    cell.font = styles['header_font']
                    cell.fill = styles['header_fill']
                    cell.alignment = styles['header_alignment']
                elif row > header_rows:
                    if (row - header_rows) % 2 == 0:
                        cell.fill = styles['gray_fill']
                    else:
                        cell.fill = styles['white_fill']
        
        # Auto-adjust column widths
        for col in range(1, max_col + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    if cell.value:
                        cell_str = str(cell.value).replace('\n', ' ')
                        max_length = max(max_length, len(cell_str))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            adjusted_width = max(adjusted_width, 10)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze headers
        if max_row > header_rows:
            ws.freeze_panes = f'A{header_rows + 1}'
        
        print(f"  Created {sheet_name}: {max_row} rows × {max_col} columns")
    
    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Save workbook
    try:
        wb.save(output_path)
        file_size = os.path.getsize(output_path)
        file_size_mb = file_size / (1024 * 1024)
        
        print(f"\n✓ Excel file created successfully!")
        print(f"  File: {output_path}")
        print(f"  Size: {file_size_mb:.2f} MB")
        print(f"  Total tables: {len(tables)}")
        
        return output_path
        
    except PermissionError:
        print(f"Error: Permission denied. Please close the Excel file: {output_path}")
        sys.exit(1)
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        sys.exit(1)


def main():
    """Main function."""
    
    if len(sys.argv) < 2:
        print("PDF to Excel Table Extractor - Smart Merged Column Detection")
        print("=" * 70)
        print("\nUsage:")
        print("  python extract_tables_smart_merged.py <pdf_file_path> [output_excel_path]")
        print("\nExamples:")
        print("  python extract_tables_smart_merged.py document.pdf")
        print("  python extract_tables_smart_merged.py document.pdf output.xlsx")
        print()
        sys.exit(1)
    
    pdf_path = sys.argv[1]
    
    if not os.path.exists(pdf_path):
        print(f"Error: File not found: {pdf_path}")
        sys.exit(1)
    
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        pdf_name = Path(pdf_path).stem
        output_path = f"/home/z/my-project/download/{pdf_name}_Tables_Smart.xlsx"
    
    if not output_path.lower().endswith('.xlsx'):
        output_path += '.xlsx'
    
    print("=" * 70)
    print("PDF to Excel Table Extractor - Smart Merged Column Detection")
    print("=" * 70)
    print()
    
    tables = extract_tables_from_pdf(pdf_path)
    
    if not tables:
        print("\nWarning: No tables found in the PDF file.")
        sys.exit(0)
    
    print(f"\nTotal tables found: {len(tables)}")
    print()
    
    create_excel_from_tables(tables, output_path)
    
    print()
    print("=" * 70)
    print("Process completed successfully!")
    print("=" * 70)
    print("\nThis version features:")
    print("  ✓ Smart detection of merged column headers")
    print("  ✓ Automatic cell merging in Excel")
    print("  ✓ Preserves hierarchical column structure")
    print("  ✓ Proper header row formatting")


if __name__ == "__main__":
    main()

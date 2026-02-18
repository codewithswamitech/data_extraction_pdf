#!/usr/bin/env python3
"""
PDF to Excel Table Extractor — img2table Engine

Uses OpenCV-based visual detection to identify table cells from the
rendered PDF image, then reads text from the native PDF layer.

This gives exact layout preservation including merged cells,
hierarchical headers, and proper column separation — without
any per-file tuning.

Usage:
    python extract_tables_img2table.py <pdf_file> [output_excel]

Examples:
    python extract_tables_img2table.py document.pdf
    python extract_tables_img2table.py document.pdf output.xlsx
"""

import sys
import os
import time
from pathlib import Path

try:
    from img2table.document import PDF
except ImportError:
    print("Error: img2table is not installed.")
    print("  pip install img2table")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl is not installed.")
    print("  pip install openpyxl")
    sys.exit(1)


# ── Styling constants ─────────────────────────────────────────────
HEADER_BG   = '1F4E79'
HEADER_FG   = 'FFFFFF'
EVEN_ROW_BG = 'FFFFFF'
ODD_ROW_BG  = 'F5F5F5'
BORDER_CLR  = '000000'


def apply_styling(xlsx_path):
    """
    Apply professional styling to an img2table-generated xlsx file.
    img2table creates the structure; we add visual polish.
    """
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path)

    header_font = Font(name='Times New Roman', size=11, bold=True, color=HEADER_FG)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(name='Times New Roman', size=10, color='000000')
    cell_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    white_fill = PatternFill(start_color=EVEN_ROW_BG, end_color=EVEN_ROW_BG, fill_type='solid')
    gray_fill  = PatternFill(start_color=ODD_ROW_BG,  end_color=ODD_ROW_BG,  fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color=BORDER_CLR),
        right=Side(style='thin', color=BORDER_CLR),
        top=Side(style='thin', color=BORDER_CLR),
        bottom=Side(style='thin', color=BORDER_CLR),
    )

    for ws in wb.worksheets:
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                # Skip merged cells — they inherit style from their parent
                if isinstance(cell, MergedCell):
                    continue
                cell.border = thin_border

                if row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                else:
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.fill = gray_fill if row % 2 == 0 else white_fill

        # Auto-width columns
        for col in range(1, max_col + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val:
                    max_len = max(max_len, len(str(val).replace('\n', ' ')))
            ws.column_dimensions[col_letter].width = max(min(max_len + 2, 50), 10)

        # Freeze first row
        if max_row > 1:
            ws.freeze_panes = 'A2'

    wb.save(xlsx_path)


def extract_pdf(pdf_path, output_path):
    """
    Extract all tables from a PDF using img2table's visual detection.
    """
    print(f"Reading PDF: {pdf_path}")

    # ── Instantiate PDF with native text extraction (no OCR needed) ──
    pdf = PDF(
        src=str(pdf_path),
        detect_rotation=False,
        pdf_text_extraction=True,   # Read text from PDF layer, skip OCR
    )

    total_pages = len(pdf.images) if hasattr(pdf, 'images') else '?'
    print(f"Total pages: {total_pages}")
    print()

    # ── Extract tables and write directly to xlsx ──
    # img2table's to_xlsx preserves exact layout including merged cells
    print("Extracting tables with visual detection (OpenCV)...")
    print("This may take a moment for large PDFs...")
    print()

    start = time.time()

    pdf.to_xlsx(
        dest=str(output_path),
        ocr=None,                   # No OCR — use native PDF text
        implicit_rows=True,         # Detect rows even without horizontal lines
        implicit_columns=True,      # Detect columns even without vertical lines
        borderless_tables=False,    # Only extract bordered tables
        min_confidence=50,
    )

    elapsed = time.time() - start

    # ── Count results ──
    if os.path.exists(output_path):
        wb = load_workbook(output_path, read_only=True)
        num_tables = len(wb.worksheets)
        wb.close()
    else:
        print("Error: Output file was not created.")
        sys.exit(1)

    print(f"  Extracted {num_tables} tables in {elapsed:.1f}s")
    print()

    # ── Apply professional styling ──
    print("Applying styling...")
    apply_styling(str(output_path))

    file_size = os.path.getsize(output_path) / (1024 * 1024)
    print()
    print("=" * 70)
    print(f"✓ Excel file created successfully!")
    print(f"  File:   {output_path}")
    print(f"  Size:   {file_size:.2f} MB")
    print(f"  Tables: {num_tables}")
    print(f"  Time:   {elapsed:.1f}s")
    print("=" * 70)
    print()
    print("Features:")
    print("  ✓ Visual cell detection (OpenCV)")
    print("  ✓ Exact layout preservation")
    print("  ✓ Merged cell support")
    print("  ✓ Native PDF text (no OCR)")
    print("  ✓ Implicit row/column detection")


def main():
    if len(sys.argv) < 2:
        print("PDF to Excel Table Extractor — img2table Engine")
        print("=" * 60)
        print()
        print("Usage:")
        print("  python extract_tables_img2table.py <pdf_file> [output.xlsx]")
        print()
        print("Examples:")
        print("  python extract_tables_img2table.py report.pdf")
        print("  python extract_tables_img2table.py report.pdf output.xlsx")
        sys.exit(1)

    pdf_path = sys.argv[1]

    if not os.path.exists(pdf_path):
        print(f"Error: File not found: {pdf_path}")
        sys.exit(1)

    # Output path
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        pdf_name = Path(pdf_path).stem
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Output_excel')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{pdf_name}_img2table.xlsx")

    if not output_path.lower().endswith('.xlsx'):
        output_path += '.xlsx'

    print("=" * 70)
    print("PDF to Excel — img2table Engine (Visual Detection)")
    print("=" * 70)
    print()

    extract_pdf(pdf_path, output_path)


if __name__ == "__main__":
    main()

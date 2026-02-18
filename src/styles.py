
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Color scheme constants
HEADER_BG_COLOR = '1F4E79'  # Dark blue
HEADER_TEXT_COLOR = 'FFFFFF'  # White
ROW_EVEN_COLOR = 'FFFFFF'    # White
ROW_ODD_COLOR = 'F5F5F5'     # Light gray
BORDER_COLOR = '000000'      # Black

def create_styles():
    """Create and return style objects for Excel formatting."""
    
    # Header style
    header_font = Font(
        name='Times New Roman',
        size=11,
        bold=True,
        color=HEADER_TEXT_COLOR
    )
    header_fill = PatternFill(
        start_color=HEADER_BG_COLOR,
        end_color=HEADER_BG_COLOR,
        fill_type='solid'
    )
    header_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Cell style (data rows)
    cell_font = Font(
        name='Times New Roman',
        size=10,
        color='000000'
    )
    cell_alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True
    )
    
    # Fill colors
    white_fill = PatternFill(
        start_color=ROW_EVEN_COLOR,
        end_color=ROW_EVEN_COLOR,
        fill_type='solid'
    )
    gray_fill = PatternFill(
        start_color=ROW_ODD_COLOR,
        end_color=ROW_ODD_COLOR,
        fill_type='solid'
    )
    
    # Border style
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_alignment': header_alignment,
        'cell_font': cell_font,
        'cell_alignment': cell_alignment,
        'white_fill': white_fill,
        'gray_fill': gray_fill,
        'border': thin_border
    }


import os
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from src.styles import create_styles


def analyze_column_structure(table):
    """
    Analyze table structure to detect hierarchical/merged column headers.
    
    Detects when a header cell spans multiple columns by looking for cells
    with content that have empty neighbors in the same row but content in
    subsequent rows below those empty cells.
    
    Returns:
        dict: {
            'header_rows': int (number of header rows detected),
            'column_groups': list of dicts with row, start_col, end_col, label
        }
    """
    if not table or len(table) < 2:
        return {'header_rows': 1, 'column_groups': []}
    
    result = {
        'header_rows': 1,
        'column_groups': []
    }
    
    # Analyze first 3 rows for header patterns
    max_analysis_rows = min(3, len(table))
    
    for row_idx in range(max_analysis_rows):
        row = table[row_idx]
        if not row:
            continue
        
        col_idx = 0
        while col_idx < len(row):
            cell_value = str(row[col_idx]).strip() if row[col_idx] else ''
            
            if cell_value:
                # Check if this cell spans multiple columns by looking ahead
                span = 1
                next_col = col_idx + 1
                
                while next_col < len(row):
                    next_cell_value = str(row[next_col]).strip() if row[next_col] else ''
                    
                    if not next_cell_value:  # Empty in current row
                        # Check if there's content in rows below this empty cell
                        has_below_content = False
                        for below_row_idx in range(row_idx + 1, min(row_idx + 3, len(table))):
                            if below_row_idx < len(table):
                                below_row = table[below_row_idx]
                                if next_col < len(below_row) and below_row[next_col]:
                                    below_val = str(below_row[next_col]).strip()
                                    if below_val:
                                        has_below_content = True
                                        break
                        
                        if has_below_content:
                            span += 1
                            next_col += 1
                        else:
                            break
                    else:
                        break
                
                if span > 1:
                    # This is a merged header spanning multiple columns
                    result['column_groups'].append({
                        'row': row_idx + 1,  # 1-based for Excel
                        'start_col': col_idx + 1,
                        'end_col': col_idx + span,
                        'label': cell_value
                    })
                    # Update header row count
                    result['header_rows'] = max(result['header_rows'], row_idx + 2)
                
                col_idx = next_col
            else:
                col_idx += 1
    
    # Cap header rows at a reasonable number
    result['header_rows'] = min(result['header_rows'], 4)
    
    return result


def create_excel_from_tables(tables, output_path):
    """
    Create an Excel workbook from extracted tables with merged column support.
    
    Args:
        tables (list): List of tables with metadata
        output_path (str): Path to save the Excel file
        
    Returns:
        str: Path to the created Excel file
    """
    print(f"\nCreating Excel workbook with smart merged column detection...")
    
    # Create workbook and remove default sheet
    wb = Workbook()
    wb.remove(wb.active)
    
    # Get styles
    styles = create_styles()
    
    # Process each table
    for table_idx, table_data in enumerate(tables, start=1):
        table = table_data['table']
        page_num = table_data['page']
        
        # Create sheet name
        sheet_name = f'Table_{table_idx}'
        ws = wb.create_sheet(title=sheet_name)
        
        # Analyze column structure for merged headers
        structure = analyze_column_structure(table)
        header_rows = structure['header_rows']
        column_groups = structure['column_groups']
        
        # Write table data
        if table:
            for row_idx, row in enumerate(table, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value is not None:
                        if isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Apply formatting
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Apply merged cells for grouped column headers
        for group in column_groups:
            try:
                ws.merge_cells(
                    start_row=group['row'],
                    start_column=group['start_col'],
                    end_row=group['row'],
                    end_column=group['end_col']
                )
            except Exception:
                pass  # Skip invalid merges (e.g. overlapping ranges)
        
        # Apply styling
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                
                # Apply basic styles
                cell.font = styles['cell_font']
                cell.alignment = styles['cell_alignment']
                cell.border = styles['border']
                
                # Header row styling
                if row <= header_rows:
                    cell.font = styles['header_font']
                    cell.fill = styles['header_fill']
                    cell.alignment = styles['header_alignment']
                # Alternating row colors for data rows
                elif (row - header_rows) % 2 == 0:
                    cell.fill = styles['gray_fill']
                else:
                    cell.fill = styles['white_fill']
        
        # Auto-adjust column widths
        for col in range(1, max_col + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    if cell.value:
                        cell_str = str(cell.value).replace('\n', ' ')
                        max_length = max(max_length, len(cell_str))
                except Exception:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            adjusted_width = max(adjusted_width, 10)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header rows
        if max_row > header_rows:
            ws.freeze_panes = f'A{header_rows + 1}'
        
        merged_info = f", {len(column_groups)} merged group(s)" if column_groups else ""
        print(f"  Created {sheet_name}: {max_row} rows × {max_col} columns (from page {page_num}){merged_info}")
    
    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Save workbook
    try:
        wb.save(output_path)
        file_size = os.path.getsize(output_path)
        file_size_mb = file_size / (1024 * 1024)
        
        print(f"\n✓ Excel file created successfully!")
        print(f"  File: {output_path}")
        print(f"  Size: {file_size_mb:.2f} MB")
        print(f"  Total tables: {len(tables)}")
        
        return output_path
        
    except PermissionError:
        print(f"Error: Permission denied. Please close the Excel file if it's open: {output_path}")
        sys.exit(1)
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        sys.exit(1)
